import openpyxl
import pandas as pd

def excel_to_markdown(file_path, sheet_name=0):
    workbook = openpyxl.load_workbook('C:/Users/Brain/Desktop/DJUR/djur1/OFERTA.xlsx', data_only=True)
    sheet = workbook[workbook.sheetnames[sheet_name]] if isinstance(sheet_name, int) else workbook[sheet_name]

    tables = []
    current_table = []

    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None and cell != '' for cell in row):
            current_table.append(row)
        elif current_table:
            tables.append(current_table)
            current_table = []

    if current_table:
        tables.append(current_table)

    # Organizar as tabelas separadas por colunas vazias
    final_tables = []
    
    for table in tables:
        df = pd.DataFrame(table).dropna(how='all').reset_index(drop=True)
        columns = df.columns.tolist()

        # Detectar separadores (colunas vazias)
        start_index = 0
        while start_index < len(columns):
            if df.iloc[:, start_index].isnull().all():  # Coluna vazia
                start_index += 1
                continue
            
            end_index = start_index
            while end_index < len(columns) and not df.iloc[:, end_index].isnull().all():
                end_index += 1
            
            # Extrair a subtabela
            sub_table = df.iloc[:, start_index:end_index]
            if not sub_table.empty:
                final_tables.append(sub_table)

            start_index = end_index
    markdown_tables = []
    for i, table in enumerate(final_tables):
        markdown = table.transpose().to_markdown(index=False)
        markdown_tables.append(f"### Tabela {i + 1}\n\n{markdown}\n\n")
    return markdown_tables


print(excel_to_markdown(r'C:\Users\Brain\Desktop\DJUR\djur1\OFERTA.xlsx'))